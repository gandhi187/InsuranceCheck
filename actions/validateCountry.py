import openpyxl
import pycountry_convert as pc
import requests
import json
from datetime import datetime, timedelta

def FindXlCell(search_str, range=None):
    """
    Iterate over a given 'range' match 'cell.value' with 'search_str'

    :param search_str: String to find
    :param range: Range to iterate, defaults to whole sheet
    :return: 'None' if no match else 
              all cells from the matching Row as 'list of cell objects'
    """
    global ws
    
    if not range:
        range = ws.iter_rows() # Defaults to whole sheet

    for tupleOfCells in range:
        for cell in tupleOfCells:
            if (cell.value == search_str):
                return [_tuple[0] for _tuple in ws.iter_cols(min_row=cell.row, max_row=cell.row)]

def FindCountryCode(search_str):
    
    cellsOfFoundRow = FindXlCell(search_str.lower())
    if cellsOfFoundRow:
    # List cellsOfFoundRow is 0-based, Index of D == 3
        print("Found:, the value of Column D is {}".format(cellsOfFoundRow[2].value))
        countryCode = cellsOfFoundRow[2].value
        return countryCode
    else:
        print("Could not find '{}' in the given cell range!".format(search_str))



def GetContinentFromCountry(country_code):
    country_continent_code = pc.country_alpha2_to_continent_code(country_code)
    continents = {
    'NA': 'Nordamerika',
    'SA': 'Südamerika', 
    'AS': 'Asien',
    'OC': 'Australien',
    'AF': 'Afrika',
    'EU': 'Europa'
    }
    continent = continents[country_continent_code]
    return continent


def getCoronaInformation(country):

    countryCode = FindCountryCode(country)
    resp = requests.get('https://api.thevirustracker.com/free-api?countryTotal='+countryCode)
    if resp.status_code != 200:
    # This means something went wrong.
        raise ApiError('GET /tasks/ {}'.format(resp.status_code))
    data =  resp.json()
    return data

def getTravelWarning(country):
    countryCode = FindCountryCode(country)
    resp = requests.get('https://www.reisewarnung.net/api?country='+countryCode)
    if resp.status_code != 200:
    # This means something went wrong.
        raise ApiError('GET /tasks/ {}'.format(resp.status_code))
    data =  resp.json()
    return data

def getJsonData():

    resp = requests.get('https://api.npoint.io/6b1f7a75909bda025ff9')
    if resp.status_code != 200:
    # This means something went wrong.
        raise ApiError('GET /tasks/ {}'.format(resp.status_code))
    data =  resp.json()
    data = json.dumps(data)
    return data

wb = openpyxl.load_workbook("actions/wiki_countrys.xlsx")
ws = wb.worksheets[0] 

print(getJsonData())
